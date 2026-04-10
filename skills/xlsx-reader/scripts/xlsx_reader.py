#!/usr/bin/env python3
"""
xlsx-reader — extract the content of an Excel (.xlsx) file, rendered
as GitHub-flavored markdown tables (one per sheet).

One of the four readers in the `document-readers` plugin. Companion
to the `document-search` skill: use document-search to locate the
file, then this skill to read its content.

Usage:
    xlsx_reader.py <file.xlsx> [--sheet "Sheet1"]
                               [--max-bytes N]
                               [--format text|json]
                               [--metadata-only]

Run with --help for the full option list.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _preflight  # noqa: E402


@dataclass
class ReadResult:
    path: str
    extension: str = ".xlsx"
    kind: str = "xlsx"
    size: int = 0
    content: Optional[str] = None
    truncated: bool = False
    metadata: dict = field(default_factory=dict)
    error: Optional[str] = None
    missing_imports: list[str] = field(default_factory=list)
    install_guide: Optional[str] = None


class MissingDependency(RuntimeError):
    def __init__(self, import_name: str, feature: str):
        self.import_name = import_name
        self.feature = feature
        super().__init__(f"missing optional package '{import_name}' for {feature}")


def _render_sheet_as_markdown(ws) -> str:
    """Render an openpyxl worksheet as a GitHub-flavored markdown table."""
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    if not rows:
        return "(empty sheet)"

    # Trim trailing fully-empty columns so the table isn't huge.
    max_cols = max(
        (max((i for i, v in enumerate(r) if v), default=-1) + 1 for r in rows),
        default=0,
    )
    if max_cols == 0:
        return "(empty sheet)"
    rows = [r[:max_cols] + [""] * (max_cols - len(r)) for r in rows]

    header = rows[0]
    body = rows[1:]
    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * max_cols) + " |")
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def read_xlsx(path: Path, sheet: Optional[str]) -> tuple[str, dict]:
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise MissingDependency("openpyxl", ".xlsx reading") from e
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    targets = [sheet] if sheet else sheet_names
    parts: list[str] = []
    returned: list[str] = []
    for name in targets:
        if name not in sheet_names:
            parts.append(f"--- Sheet: {name} (not found) ---")
            continue
        ws = wb[name]
        parts.append(f"--- Sheet: {name} ---")
        parts.append(_render_sheet_as_markdown(ws))
        returned.append(name)
    metadata = {
        "sheets": sheet_names,
        "sheets_returned": returned,
    }
    return "\n".join(parts), metadata


def read_file(
    path: Path,
    *,
    sheet: Optional[str],
    max_bytes: int,
    metadata_only: bool,
) -> ReadResult:
    result = ReadResult(path=str(path), size=path.stat().st_size)
    try:
        content, meta = read_xlsx(path, sheet)
        result.metadata = meta
    except MissingDependency as e:
        result.error = str(e)
        result.missing_imports = [e.import_name]
        result.install_guide = _preflight.format_install_guide(
            [e.import_name],
            feature="reading .xlsx files",
            skill_name="xlsx-reader",
        )
        return result
    except Exception as e:  # noqa: BLE001
        result.error = f"read failed: {e}"
        return result

    if content is not None and len(content.encode("utf-8", errors="ignore")) > max_bytes:
        content = content.encode("utf-8", errors="ignore")[:max_bytes].decode("utf-8", errors="replace")
        result.truncated = True

    if not metadata_only:
        result.content = content
    return result


def render_text(r: ReadResult) -> str:
    out: list[str] = []
    out.append("# xlsx-reader")
    out.append(f"path:      {r.path}")
    out.append(f"size:      {r.size} bytes")
    if r.truncated:
        out.append("truncated: yes (output capped by --max-bytes)")
    if r.metadata:
        out.append("metadata:")
        for k, v in r.metadata.items():
            out.append(f"  {k}: {v}")
    out.append("")

    if r.error:
        out.append(f"error: {r.error}")
        if r.install_guide:
            out.append("")
            out.append("## Install guide")
            out.append(r.install_guide)
        return "\n".join(out)

    if r.content is None:
        out.append("(metadata-only mode; no content returned)")
        return "\n".join(out)

    out.append("## Content")
    out.append(r.content)
    return "\n".join(out)


def render_json(r: ReadResult) -> str:
    return json.dumps(asdict(r), ensure_ascii=False, indent=2)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsx_reader.py",
        description="Extract the content of an Excel (.xlsx) file, "
                    "rendered as GitHub-flavored markdown tables.",
    )
    p.add_argument("file", help="Path to the .xlsx file to read.")
    p.add_argument("--sheet", default=None,
                   help="Sheet name to extract. If omitted, all sheets are returned.")
    p.add_argument("--max-bytes", type=int, default=200_000,
                   help="Maximum bytes of content returned (default: 200000).")
    p.add_argument("--format", choices=("text", "json"), default="text",
                   help="Output format (default: text).")
    p.add_argument("--metadata-only", action="store_true",
                   help="Return only metadata (sheet names) without the content.")
    return p


def main(argv: Optional[list[str]] = None) -> int:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        except (AttributeError, ValueError):
            pass

    args = build_parser().parse_args(argv)
    path = Path(args.file)
    if not path.exists():
        print(f"error: file does not exist: {path}", file=sys.stderr)
        return 2
    if not path.is_file():
        print(f"error: not a regular file: {path}", file=sys.stderr)
        return 2
    if path.suffix.lower() != ".xlsx":
        print(f"warning: file does not have a .xlsx extension: {path}", file=sys.stderr)

    result = read_file(
        path,
        sheet=args.sheet,
        max_bytes=args.max_bytes,
        metadata_only=args.metadata_only,
    )
    if args.format == "json":
        print(render_json(result))
    else:
        print(render_text(result))
    return 0 if result.error is None else 2


if __name__ == "__main__":
    sys.exit(main())
