#!/usr/bin/env python3
"""
xlsx-inspector — structural inspection of an Excel (.xlsx) workbook.

Unlike xlsx-reader, which renders sheets as Markdown tables, this
skill returns *structure* — the deterministic facts an LLM cannot
derive from a rendered table view:

  * Workbook properties (title, creator, dates, ...) plus per-sheet
    dimensions, cell counts, and merged-range counts
  * Formula dependency graph — which cells contain formulas, which
    cells/ranges those formulas reference, and which references
    cross sheet boundaries (best-effort regex, not a full parser)
  * Named range inventory, with workbook-vs-sheet scope

Sibling skills in the `document-inspector` plugin handle other formats
(pdf-inspector for .pdf, docx-inspector for .docx). This skill is
.xlsx-only by design.

Usage:
    xlsx_inspector.py <file> --feature metadata|formulas|named-ranges|all
                             [--format text|json]

Run with --help for the full option list.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Optional

# Self-contained install-guide helper.
sys.path.insert(0, str(Path(__file__).resolve().parent))
import _preflight  # noqa: E402


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class InspectResult:
    path: str
    extension: str = ".xlsx"
    kind: str = "xlsx"
    feature: str = "all"
    size: int = 0
    metadata: dict = field(default_factory=dict)
    formulas: dict = field(default_factory=dict)
    named_ranges: list = field(default_factory=list)
    error: Optional[str] = None
    missing_imports: list[str] = field(default_factory=list)
    install_guide: Optional[str] = None


class MissingDependency(RuntimeError):
    def __init__(self, import_name: str, feature: str):
        self.import_name = import_name
        self.feature = feature
        super().__init__(f"missing optional package '{import_name}' for {feature}")


# ---------------------------------------------------------------------------
# Open + dispatch
# ---------------------------------------------------------------------------

FEATURES = ("metadata", "formulas", "named-ranges", "all")


def _open_xlsx(path: Path):
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise MissingDependency("openpyxl", "xlsx inspection") from e
    # data_only=False preserves formula strings (True would return
    # cached results). read_only=False gives us full defined_names
    # access. keep_links=False avoids resolving external workbook links.
    return openpyxl.load_workbook(
        str(path), data_only=False, read_only=False, keep_links=False
    )


def _datetime_to_iso(value: Any) -> Optional[str]:
    if value is None:
        return None
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        try:
            return iso()
        except Exception:  # noqa: BLE001
            pass
    return str(value)


# ---------------------------------------------------------------------------
# Feature: metadata
# ---------------------------------------------------------------------------

_DOC_PROPS = (
    "creator", "title", "subject", "description", "keywords",
    "category", "last_modified_by", "revision", "version", "identifier",
)


def _xlsx_metadata(wb) -> dict:
    out: dict[str, Any] = {}
    props = wb.properties

    for name in _DOC_PROPS:
        value = getattr(props, name, None)
        out[name] = None if value in (None, "") else str(value)

    out["created"] = _datetime_to_iso(getattr(props, "created", None))
    out["modified"] = _datetime_to_iso(getattr(props, "modified", None))

    # Workbook-level facts.
    sheet_names = list(wb.sheetnames)
    out["sheet_count"] = len(sheet_names)

    active_idx: Optional[int] = None
    try:
        active_title = wb.active.title
        if active_title in sheet_names:
            active_idx = sheet_names.index(active_title)
    except Exception:  # noqa: BLE001
        pass
    out["active_sheet_index"] = active_idx

    sheets_info: list[dict[str, Any]] = []
    for name in sheet_names:
        ws = wb[name]
        info: dict[str, Any] = {"name": name}
        try:
            info["max_row"] = int(ws.max_row or 0)
            info["max_column"] = int(ws.max_column or 0)
        except Exception:  # noqa: BLE001
            info["max_row"] = None
            info["max_column"] = None
        try:
            info["dimensions"] = str(ws.dimensions)
        except Exception:  # noqa: BLE001
            info["dimensions"] = None
        try:
            info["merged_cells_count"] = len(list(ws.merged_cells.ranges))
        except Exception:  # noqa: BLE001
            info["merged_cells_count"] = 0
        # Count non-empty cells. Efficient on small/medium workbooks;
        # capped by max_row * max_column as openpyxl naturally does.
        try:
            data_cell_count = 0
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is not None and v != "":
                        data_cell_count += 1
            info["data_cell_count"] = data_cell_count
        except Exception:  # noqa: BLE001
            info["data_cell_count"] = None
        sheets_info.append(info)
    out["sheets"] = sheets_info
    return out


# ---------------------------------------------------------------------------
# Feature: formulas
# ---------------------------------------------------------------------------

_DQUOTED = re.compile(r'"[^"]*"')

# Cross-sheet reference: 'Sheet Name'!A1  or  SheetName!A1:B5
_CROSS_SHEET_REF = re.compile(
    r"""
    (?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))   # 1: 'Sheet Name' OR 2: SheetName
    !                                           # sheet separator
    (\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)     # 3: cell or range
    """,
    re.VERBOSE,
)

# Local reference: A1, $A$1, A1:B5 — must contain both letters and digits
# and not be adjacent to more alphabetic chars (avoid matching inside
# function names like 'SUM' or 'IF'). We enforce word boundaries and a
# required digit via a secondary check.
_LOCAL_REF = re.compile(
    r"(?<![A-Za-z_])(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)(?![A-Za-z_])"
)

_FORMULA_SAMPLES_CAP = 50


def _extract_references(formula: str) -> list[dict[str, Any]]:
    """
    Return a list of {"sheet": str|None, "range": str} references
    found in `formula`. Best-effort: strips quoted string literals
    first so A1-like text inside a literal doesn't get counted.
    """
    cleaned = _DQUOTED.sub('""', formula)
    refs: list[dict[str, Any]] = []
    seen: set[tuple[Optional[str], str]] = set()

    # Cross-sheet refs first, then remove them from the string so their
    # range portions don't get double-counted as local refs.
    cross_stripped_parts: list[str] = []
    pos = 0
    for m in _CROSS_SHEET_REF.finditer(cleaned):
        cross_stripped_parts.append(cleaned[pos:m.start()])
        sheet = m.group(1) or m.group(2)
        rng = m.group(3)
        key = (sheet, rng)
        if key not in seen:
            seen.add(key)
            refs.append({"sheet": sheet, "range": rng})
        cross_stripped_parts.append(" ")  # placeholder to keep offsets sane
        pos = m.end()
    cross_stripped_parts.append(cleaned[pos:])
    cleaned_local = "".join(cross_stripped_parts)

    for m in _LOCAL_REF.finditer(cleaned_local):
        rng = m.group(1)
        key = (None, rng)
        if key not in seen:
            seen.add(key)
            refs.append({"sheet": None, "range": rng})

    return refs


def _xlsx_formulas(wb) -> dict:
    """
    Walk every sheet and every cell, building a best-effort formula
    dependency report. Returns the shape documented in the module
    docstring + SKILL.md.
    """
    formula_count = 0
    cross_sheet_count = 0
    dependents_by_sheet: dict[str, int] = {}
    referenced_sheets: set[str] = set()
    samples: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                formula_count += 1
                sheet_formula_count += 1
                refs = _extract_references(value)
                has_cross = any(
                    r["sheet"] is not None and r["sheet"] != sheet_name
                    for r in refs
                )
                if has_cross:
                    cross_sheet_count += 1
                for r in refs:
                    if r["sheet"] is not None:
                        referenced_sheets.add(r["sheet"])
                if len(samples) < _FORMULA_SAMPLES_CAP:
                    samples.append({
                        "cell": f"{sheet_name}!{cell.coordinate}",
                        "formula": value,
                        "references": refs,
                        "cross_sheet": has_cross,
                    })
        dependents_by_sheet[sheet_name] = sheet_formula_count

    return {
        "formula_cell_count": formula_count,
        "cross_sheet_reference_count": cross_sheet_count,
        "dependents_by_sheet": dependents_by_sheet,
        "unique_referenced_sheets": sorted(referenced_sheets),
        "sample_formulas": samples,
        "truncated": formula_count > _FORMULA_SAMPLES_CAP,
    }


# ---------------------------------------------------------------------------
# Feature: named-ranges
# ---------------------------------------------------------------------------

def _xlsx_named_ranges(wb) -> list[dict[str, Any]]:
    """
    Iterate wb.defined_names, tolerating API differences across
    openpyxl versions. Returns a list of {name, value, scope}.
    """
    result: list[dict[str, Any]] = []
    sheetnames = list(wb.sheetnames)

    # openpyxl >= 3.1 — defined_names is DefinedNameDict, dict-like.
    try:
        iter_names = list(wb.defined_names)
    except TypeError:
        iter_names = []
    for name in iter_names:
        try:
            defn = wb.defined_names[name]
        except Exception:  # noqa: BLE001
            continue
        local_id = getattr(defn, "localSheetId", None)
        if local_id is None:
            scope = "workbook"
        else:
            try:
                scope = sheetnames[int(local_id)]
            except Exception:  # noqa: BLE001
                scope = f"sheet[{local_id}]"
        value = (
            getattr(defn, "value", None)
            or getattr(defn, "attr_text", None)
            or ""
        )
        result.append({
            "name": str(name),
            "value": str(value),
            "scope": scope,
        })

    # Older openpyxl fallback — defined_names.definedName.
    if not result:
        try:
            for defn in wb.defined_names.definedName:  # type: ignore[attr-defined]
                local_id = getattr(defn, "localSheetId", None)
                if local_id is None:
                    scope = "workbook"
                else:
                    try:
                        scope = sheetnames[int(local_id)]
                    except Exception:  # noqa: BLE001
                        scope = f"sheet[{local_id}]"
                result.append({
                    "name": str(defn.name),
                    "value": str(getattr(defn, "value", "")),
                    "scope": scope,
                })
        except Exception:  # noqa: BLE001
            pass

    return result


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def inspect(path: Path, feature: str) -> InspectResult:
    result = InspectResult(
        path=str(path),
        extension=path.suffix.lower(),
        kind="xlsx" if path.suffix.lower() == ".xlsx" else "unknown",
        feature=feature,
        size=path.stat().st_size,
    )

    if result.extension != ".xlsx":
        result.error = (
            f"xlsx-inspector only accepts .xlsx files "
            f"(got {result.extension or '(none)'}) — use pdf-inspector "
            f"or docx-inspector for other formats"
        )
        return result

    try:
        wb = _open_xlsx(path)
    except MissingDependency as e:
        result.error = str(e)
        result.missing_imports = [e.import_name]
        result.install_guide = _preflight.format_install_guide(
            [e.import_name],
            feature="inspecting .xlsx files",
            skill_name="xlsx-inspector",
        )
        return result
    except Exception as e:  # noqa: BLE001
        result.error = f"failed to open XLSX: {e}"
        return result

    try:
        if feature in ("metadata", "all"):
            result.metadata = _xlsx_metadata(wb)
        if feature in ("formulas", "all"):
            result.formulas = _xlsx_formulas(wb)
        if feature in ("named-ranges", "all"):
            result.named_ranges = _xlsx_named_ranges(wb)
    except Exception as e:  # noqa: BLE001
        result.error = f"inspection failed: {e}"
        return result
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass

    return result


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def render_text(r: InspectResult) -> str:
    out: list[str] = []
    out.append("# xlsx-inspector")
    out.append(f"path:      {r.path}")
    out.append(f"size:      {r.size} bytes")
    out.append(f"kind:      {r.kind}")
    out.append(f"feature:   {r.feature}")
    out.append("")

    if r.error:
        out.append(f"error: {r.error}")
        if r.install_guide:
            out.append("")
            out.append("## Install guide")
            out.append(r.install_guide)
        return "\n".join(out)

    if r.feature in ("metadata", "all") and r.metadata:
        out.append("## Metadata")
        ordered_keys = (
            "title", "creator", "subject", "description", "keywords",
            "category", "last_modified_by", "revision", "version", "identifier",
            "created", "modified",
            "sheet_count", "active_sheet_index",
        )
        for key in ordered_keys:
            if key in r.metadata:
                val = r.metadata[key]
                out.append(f"  {key}: {val if val is not None else '(none)'}")
        sheets = r.metadata.get("sheets") or []
        if sheets:
            out.append("")
            out.append("  sheets:")
            for s in sheets:
                out.append(
                    f"    - {s.get('name')}: dim={s.get('dimensions')}, "
                    f"rows={s.get('max_row')}, cols={s.get('max_column')}, "
                    f"cells={s.get('data_cell_count')}, "
                    f"merged={s.get('merged_cells_count')}"
                )
        out.append("")

    if r.feature in ("formulas", "all") and r.formulas:
        out.append("## Formulas")
        out.append(f"  formula_cell_count:          {r.formulas.get('formula_cell_count', 0)}")
        out.append(f"  cross_sheet_reference_count: {r.formulas.get('cross_sheet_reference_count', 0)}")
        ref_sheets = r.formulas.get("unique_referenced_sheets") or []
        if ref_sheets:
            out.append(f"  unique_referenced_sheets:    {', '.join(ref_sheets)}")
        deps = r.formulas.get("dependents_by_sheet") or {}
        if deps:
            out.append("  dependents_by_sheet:")
            for sheet, count in deps.items():
                out.append(f"    {sheet}: {count}")
        samples = r.formulas.get("sample_formulas") or []
        if samples:
            out.append("")
            out.append("  sample_formulas:")
            for s in samples[:30]:
                out.append(f"    {s.get('cell')}: {s.get('formula')}")
                refs = s.get("references") or []
                if refs:
                    parts = [
                        f"{r_.get('sheet') + '!' if r_.get('sheet') else ''}{r_.get('range')}"
                        for r_ in refs
                    ]
                    out.append(f"      refs: {', '.join(parts)}")
            if r.formulas.get("truncated"):
                out.append(f"    ... (more; capped at {_FORMULA_SAMPLES_CAP} samples)")
        out.append("")

    if r.feature in ("named-ranges", "all") and r.named_ranges:
        out.append("## Named ranges")
        for nr in r.named_ranges[:100]:
            out.append(f"  [{nr.get('scope')}] {nr.get('name')} = {nr.get('value')}")
        if len(r.named_ranges) > 100:
            out.append(f"  ... ({len(r.named_ranges) - 100} more)")
        out.append("")
    elif r.feature in ("named-ranges", "all"):
        out.append("## Named ranges")
        out.append("  (none)")
        out.append("")

    return "\n".join(out).rstrip() + "\n"


def render_json(r: InspectResult) -> str:
    return json.dumps(asdict(r), ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsx_inspector.py",
        description="Structural inspection of a single Excel (.xlsx) "
                    "workbook — metadata, formula dependency graph, and "
                    "named-range inventory. Returns deterministic facts "
                    "an LLM cannot derive from a rendered table view.",
    )
    p.add_argument("file", help="Path to the .xlsx file to inspect.")
    p.add_argument("--feature", choices=FEATURES, default="all",
                   help="Which inspection to run (default: all).")
    p.add_argument("--format", choices=("text", "json"), default="text",
                   help="Output format (default: text).")
    return p


def main(argv: Optional[list[str]] = None) -> int:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        except (AttributeError, ValueError):
            pass

    args = build_parser().parse_args(argv)
    path = Path(args.file)
    if not path.exists():
        print(f"error: file does not exist: {path}", file=sys.stderr)
        return 2
    if not path.is_file():
        print(f"error: not a regular file: {path}", file=sys.stderr)
        return 2

    result = inspect(path, feature=args.feature)

    if args.format == "json":
        print(render_json(result))
    else:
        print(render_text(result))
    return 0 if result.error is None else 2


if __name__ == "__main__":
    sys.exit(main())
